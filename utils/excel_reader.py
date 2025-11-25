import os
from openpyxl import load_workbook

def get_excel_data(sheet_name: str, file_name: str = "testdata_login.xlsx"):
    root_dir = os.path.dirname(os.path.dirname(__file__))
    file_path = os.path.join(root_dir, "resources", file_name)

    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    data = []
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    return data
